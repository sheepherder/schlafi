#!/usr/bin/env python3
"""
build_wordlist.py – Generiert words.js fuer Schlummershuffle.

Quellen:
  1. Wortfrequenz: hermitdave/FrequencyWords (OpenSubtitles-Korpus)
     https://github.com/hermitdave/FrequencyWords
  2. Substantive + Genus: gambolputty/german-nouns (Wiktionary-Extrakt)
     https://github.com/gambolputty/german-nouns
  3. Psycholinguistische Bewertungen: German Psycholinguistic Toolbox
     Brysbaert et al., https://osf.io/ghjd2/
     Spalten: GPT_concr (Konkretheit), GPT_val (Valenz), GPT_arou (Arousal)

Methodik:
  - Frequenzliste mit ~1M deutschen Woertern (OpenSubtitles)
  - Kreuzreferenz mit Wiktionary-Substantiven (nur pos == "Substantiv" exakt)
  - Nur Lemma-Formen (keine Plurale, Genitive, flektierte Formen)
  - Filter: Konkretheit >= 5.8 (gut vorstellbar), Skala 1-7
  - Filter: Valenz >= 3.5 (nicht negativ/bedrohlich), Skala 1-7
  - Filter: Arousal <= 4.0 (beruhigend, nicht aufregend), Skala 1-7
  - Ausschluss nominalisierter Infinitive und falscher Substantive
  - Sortiert nach Haeufigkeit, Top ~3400 Eintraege

Abhaengigkeiten: openpyxl (pip install openpyxl)

Nutzung: python3 scripts/build_wordlist.py
"""

import csv
import io
import os
import sys
import urllib.request

FREQ_URL = "https://raw.githubusercontent.com/hermitdave/FrequencyWords/master/content/2018/de/de_full.txt"
NOUNS_URL = "https://raw.githubusercontent.com/gambolputty/german-nouns/master/german_nouns/nouns.csv"
CONCRETENESS_URL = "https://osf.io/download/qftv9/"

TARGET_COUNT = 3500
CONCRETENESS_MIN = 5.8   # Skala 1-7, >= 5.8 = gut vorstellbar
VALENCE_MIN = 3.5         # Skala 1-7, >= 3.5 = nicht negativ
AROUSAL_MAX = 4.0         # Skala 1-7, <= 4 = beruhigend

# ── Blocklist: Woerter die nicht in die Liste gehoeren ──

BLOCKED = {
    # Primaer keine Substantive (Farben, Adverbien, Zahlwoerter)
    'gelb', 'orange', 'morgen', 'sekunde', 'zehn', 'zwölf',
    # Plural-Lemmata
    'schiffe', 'taxis', 'socken', 'hosenträger', 'boxershorts', 'hufe',
    'anchovis', 'wollsocken',
    # Veraltet / dialektal / ungebraeuchlich
    'ross', 'gaul', 'tornister', 'kaute', 'krümme', 'fingerbreit',
    'puschen', 'schlappen', 'nuckel', 'wauwau', 'dreikäsehoch',
    'schneemensch', 'ozeandampfer',
    # Kurzformen / Varianten (Duplikat-Vermeidung, weniger gebraeuchlich behalten)
    'katz', 'mieze', 'miezekatze', 'pulli', 'türe', 'limo',
    # Lautmalerei / kein echtes Substantiv
    'miau', 'pups', 'rülpser', 'schmatz', 'nieser',
    # Nominalisierte Verben auf -eln/-ern
    'pinkeln',
    # Eigennamen / Marken / Orte (nicht vom Wiktionary-Filter erfasst)
    'pazifik', 'pazifikküste', 'vatikanstadt', 'mosel', 'klagemauer',
    'davidstern', 'nordpol', 'südpol',
    'nissan', 'volkswagen', 'ipod', 'walkman', 'blackberry', 'tesafilm',
    'trabi',
    # Penny / Waehrungen
    'penny',
    # Lake (Salzlake)
    'lake',
    # Abstrakte Begriffe / Nicht visualisierbar
    'beweisstück', 'fluchtweg', 'atemzug', 'bissen', 'handschlag',
    'händedruck', 'händeschütteln', 'knopfdruck', 'wimpernschlag',
    'hahnenschrei', 'stuhlgang', 'kopfstand', 'liegestütz', 'purzelbaum',
    'spatenstich', 'boxenstopp', 'barzahlung', 'siedepunkt', 'brustumfang',
    'rückwärtsgang', 'angelausflug', 'schlittenfahrt', 'taxifahrt',
    'friseurtermin', 'reifenwechsel', 'sonnabend', 'sommertag',
    'kissenschlacht',
    # Abstrakte Nummern / Dokumente / Masseinheiten
    'telefonnummer', 'handynummer', 'zimmernummer', 'kontonummer',
    'hausnummer', 'postleitzahl', 'steuernummer', 'privatadresse',
    'geburtsdatum', 'geburtsurkunde', 'heiratsurkunde', 'trauschein',
    'kontoauszug', 'sparkonto', 'notgroschen', 'stimmzettel',
    'wahlzettel', 'spickzettel', 'einkaufszettel',
    'stundenkilometer', 'quadratkilometer', 'quadratzentimeter',
    'kubikmeter', 'milligramm', 'kilogramm', 'millisekunde',
    'nanosekunde', 'zentner', 'meile', 'pfund',
    # Unangenehm / unpassend fuer Schlaf-App
    'gesäß', 'pobacke', 'popo', 'bierbauch', 'hosenstall',
    'spaghettifresser', 'glatzkopf', 'tölpel',
    # Bargeld-Begriffe (unzaehlbar / abstrakt)
    'bargeld', 'kleingeld', 'wechselgeld',
    # Technische Abkuerzungen
    'cpu', 'lkw', 'pixel', 'screenshot', 'desktop',
    # Bolognese (Adjektiv als Substantiv)
    'bolognese',
    # Veraltetete Technik
    'autotelefon', 'wählscheibe',
}

# ── Unzaehlbare Substantive / Stoffnamen ──
# Einzelwoerter die unnatuerlich mit "ein/eine" klingen.
# Komposita werden zusaetzlich per Suffix-Muster erkannt (siehe UNCOUNTABLE_SUFFIXES).
UNCOUNTABLE = {
    # Fluessigkeiten / Getraenke
    'wasser', 'milch', 'sahne', 'rahm', 'brühe', 'bouillon',
    'tee', 'kaffee', 'saft', 'sirup', 'diesel',
    # Lebensmittel (unzaehlbar)
    'eis', 'brot', 'reis', 'mais', 'weizen', 'hafer', 'roggen',
    'gerste', 'hirse', 'grieß', 'buchweizen',
    'fleisch', 'obst', 'gemüse', 'tofu', 'popcorn', 'sauerkraut',
    'spinat', 'spargel', 'brokkoli', 'broccoli', 'rosenkohl',
    'grünkohl', 'rotkohl', 'weißkohl', 'wirsing', 'blaukraut',
    'rhabarber', 'meerrettich', 'hopfen',
    'spaghetti', 'pasta', 'risotto', 'sushi',
    'kartoffelbrei', 'kartoffelpüree', 'haferbrei', 'haferschleim',
    'grießbrei', 'milchreis', 'apfelmus', 'kompott',
    'eiscreme', 'eiskrem',
    # Gewuerze / Kraeuter
    'knoblauch', 'pfeffer', 'senf', 'zimt', 'vanille', 'ingwer',
    'rosmarin', 'thymian', 'oregano', 'koriander', 'safran',
    'muskat', 'sesam', 'kümmel', 'kreuzkümmel', 'schnittlauch',
    'petersilie', 'basilikum', 'lavendel', 'baldrian', 'fenchel',
    # Kaesesorten / Aufstriche / Sossen
    'mozzarella', 'parmesan', 'ricotta', 'feta', 'frischkäse',
    'hüttenkäse', 'schmelzkäse', 'ziegenkäse',
    'mayo', 'mayonnaise', 'guacamole', 'margarine', 'marmelade',
    'konfitüre', 'erdnussbutter', 'marzipan', 'lakritze',
    # Materialien / Stoffe
    'holz', 'sand', 'gras', 'heu', 'stroh', 'wolle', 'seide',
    'baumwolle', 'leder', 'gummi', 'porzellan', 'moos', 'lehm',
    'marmor', 'granit', 'asphalt', 'kies', 'schotter',
    'bambus', 'kautschuk', 'elfenbein', 'perlmutt', 'mahagoni',
    'jute', 'linoleum', 'parkett',
    # Metalle / Elemente / Chemie
    'gold', 'silber', 'kupfer', 'eisen', 'stahl', 'blei', 'zinn',
    'zink', 'nickel', 'messing', 'bronze', 'aluminium', 'platin',
    'helium', 'magnesium', 'kobalt', 'cobalt', 'silizium', 'silicium',
    'cäsium', 'graphit', 'obsidian',
    # Baustoffe / Klebstoffe
    'beton', 'zement', 'wachs', 'mörtel', 'silikon', 'styropor',
    'teflon', 'zellophan', 'leim', 'klebstoff', 'kleber',
    # Sonstige Stoffe
    'öl', 'salz', 'zucker', 'butter', 'honig', 'mehl',
    'staub', 'schnee', 'laub', 'regen', 'nebel', 'rauch',
    'ketchup', 'essig', 'sahne',
    'papier', 'blech', 'erz', 'garn', 'watte', 'pappe',
    'tinte', 'kreide', 'kohle', 'knete', 'hefe', 'mulch', 'humus',
    'salpeter',
    # Medizin-Stoffe
    'insulin', 'penicillin', 'penizillin', 'paracetamol', 'vaseline',
    # Sonnenlicht / Sonnenschein (unzaehlbar)
    'sonnenschein', 'sonnenlicht', 'kerzenlicht', 'kerzenschein',
    'frischluft', 'atemluft',
    # Geschirr/Besteck (Kollektivnomen)
    'geschirr', 'besteck', 'winterkleidung',
    # Weitere Stoffe / Materialien
    'shampoo', 'zahnpasta', 'zahnseide', 'pomade', 'autopolitur',
    'schaumstoff', 'erdgas', 'propan', 'trockeneis',
    'wellblech', 'laminat', 'lötzinn', 'terrakotta', 'alabaster',
    'dachpappe', 'teak',
    'natron', 'natriumchlorid', 'harnstoff', 'iod',
    'glukose', 'glucose', 'dextrose', 'zitronensäure', 'melasse',
    'mett', 'sauerrahm', 'kurkuma', 'pflaumenmus', 'kaffeesatz',
    'babynahrung',
    'wildleder', 'schweinsleder', 'kalbsleder',
    'packeis', 'reisig', 'strandgut', 'ackerland',
    'eisenerz', 'felsgestein', 'wüstensand',
    'desinfektionsmittel', 'waschmittel', 'spülmittel',
    'schmierseife', 'flüssigseife',
    'blutplasma', 'zahnschmelz',
}

# Suffixe fuer Komposita die fast immer unzaehlbar / Stoffnamen sind
UNCOUNTABLE_SUFFIXES = (
    'wasser', 'milch', 'öl', 'mehl', 'salz', 'holz', 'fleisch',
    'creme', 'sahne', 'butter', 'pulver', 'staub', 'zucker',
    'soße', 'sauce', 'sirup', 'milch', 'glas',  # Panzerglas, Fiberglas
    'gold', 'stahl', 'beton',
    'wachs', 'kohle',  # Holzkohle, Kerzenwachs
    'schnee', 'laub', 'gras',  # Seegras, Herbstlaub
    'kraut',  # Heidekraut, Sauerkraut
    'mark',   # Tomatenmark (but not Knochenmark which is medical anyway)
    'reis',   # Milchreis
    'kohl',   # Rosenkohl, Blumenkohl (uncountable vegetables)
    'tang',   # Seetang
    'papier', # Geschenkpapier, Kohlepapier, etc.
    'leder',  # Wildleder, Kalbsleder, etc.
    'sand',   # Wüstensand, Vogelsand
    'wolle',  # Holzwolle, Stahlwolle
    'mittel', # Waschmittel, Spülmittel, Desinfektionsmittel
    'seife',  # Schmierseife, Flüssigseife
    'politur', # Autopolitur, Möbelpolitur
)

# Suffixe fuer medizinische / anatomische Begriffe
MEDICAL_SUFFIXES = (
    'drüse', 'nerv', 'höhle', 'röhre', 'gewebe', 'muskel',
    'klappe', 'sehne', 'scheibe',  # Bandscheibe
    'wirbel',  # Lendenwirbel
    'darm',    # Blinddarm, Dickdarm
    'hirn',    # Kleinhirn
    'mark',    # Knochenmark, Rückenmark
    'flügel',  # Lungenflügel
    'haut',    # Kopfhaut, Nagelhaut, Netzhaut, Vorhaut
)

# Einzelne medizinische / intime Begriffe
MEDICAL = {
    'gebärmutter', 'uterus', 'prostata', 'plazenta', 'fruchtblase',
    'eizelle', 'vorhaut', 'harnröhre', 'eileiter', 'eierstock',
    'muttermund', 'nabelschnur', 'schambein',
    'blinddarm', 'milz', 'leber', 'lunge', 'niere', 'magen',
    'gallenblase', 'bauchspeicheldrüse', 'pankreas',
    'dickdarm', 'dünndarm', 'lymphknoten', 'kehlkopf',
    'arterie', 'aorta', 'halsschlagader', 'speiseröhre', 'luftröhre',
    'wirbelsäule', 'knochenmark', 'rückenmark',
    'schlüsselbein', 'brustbein', 'wadenbein', 'bandscheibe',
    'achillessehne', 'retina', 'netzhaut', 'pupille',
    'augapfel', 'augenhöhle', 'bauchhöhle', 'achselhöhle',
    'gurgel', 'kehle', 'schläfe', 'brustkorb', 'zwerchfell',
    'knorpel', 'bizeps', 'trizeps',
    'herzmuskel', 'lungenflügel', 'lungengewebe', 'nebenniere',
    'herzklappe', 'schließmuskel', 'lendenwirbel', 'gehörgang',
    'kopfhaut', 'nagelhaut', 'trommelfell', 'tränendrüse',
    'kleinhirn', 'sehnerv',
    # Medizinische Gegenstaende
    'tampon', 'antibabypille', 'kanüle', 'druckverband',
    'kochsalzlösung', 'tetanusimpfung', 'herzschrittmacher',
    'stethoskop', 'geigerzähler',
}

# Duplikate: bei Mehrfacheintraegen fuer dasselbe Konzept wird nur der
# haeufigste behalten. Diese Liste enthaelt die WENIGER gebraeuchlichen.
DUPLICATE_REMOVE = {
    # Schreibvarianten (seltenere entfernen)
    'delphin', 'penizillin', 'grammofon', 'cobalt', 'silicium',
    'videorecorder', 'portmonee', 'eiskrem', 'broccoli', 'wagon',
    'omelette', 'kräcker',
    # Konzept-Duplikate (seltenere/informellere entfernen)
    'piano', 'violine', 'fiedel',           # -> Klavier, Geige
    'flusspferd', 'hippo',                   # -> Nilpferd
    'rhinozeros',                            # -> Nashorn
    'briefträger', 'postmann',               # -> Postbote
    'ellenbogen',                            # -> Ellbogen
    'yachthafen',                            # -> Jachthafen
    'armreifen',                             # -> Armreif
    'schubkarren',                           # -> Schubkarre
    'pedale',                                # -> Pedal
    'friseuse', 'friseursalon', 'friseurladen', 'frisiersalon', 'frisörsalon',
    'tresorraum', 'geldschrank', 'panzerschrank',  # -> Tresor
    'kerzenständer', 'kerzenleuchter',       # -> Kerzenhalter
    'sparbüchse', 'spardose',               # -> Sparschwein
    'zahncreme',                             # -> Zahnpasta
    'geldbörse', 'portemonnaie',            # -> Brieftasche
    'tomatensoße', 'tomatensauce',          # -> (both sauce, keep neither? keep Tomatensoße)
    'sojasauce',                            # -> Sojasoße
    'polizeiwache', 'polizeipräsidium', 'polizeiposten',  # -> Polizeirevier
    'kartoffelpüree',                       # -> Kartoffelbrei
    'haferschleim',                         # -> Haferbrei
    'knetmasse',                            # -> Knete
    # Weitere Duplikate
    'fernsehapparat', 'farbfernseher', 'fernsehgerät',  # -> Fernseher
    'omi', 'großmama',                     # -> Oma
    'teddybär',                            # -> Teddy (kuerzer, gebraeuchlicher)
    'pandabär',                            # -> Panda
    'walfisch',                            # -> Wal
    'automobil',                           # -> Auto
    'autobus',                             # -> Bus
    'seemöwe',                             # -> Möwe
    'mikrophon',                           # -> Mikrofon
    'babyphon',                            # -> Babyfon
    'lolli', 'dauerlutscher',              # -> Lutscher
    'leihwagen',                           # -> Mietwagen
    'autowaschanlage', 'waschstraße',      # -> Waschanlage
    'mohrrübe', 'karotte',                 # -> Möhre
    'busstation',                          # -> Bushaltestelle
    'wagenschlüssel', 'zündschlüssel',     # -> Autoschlüssel
    'enkeltochter',                        # -> Enkelin
    'bulette', 'boulette',                 # -> Frikadelle
    'limone',                              # -> Limette
    'stiege',                              # -> Treppe
    'truck',                               # -> Lastwagen
}


# Genus-Korrekturen fuer Woerter deren erster Wiktionary-Eintrag falsch ist
GENUS_OVERRIDE = {
    'foto': 'n',       # das Foto, nicht der Foto
    'tram': 'f',       # die Tram
    'kiwi': 'f',       # die Kiwi (Frucht), nicht der Kiwi (Vogel)
}


def is_uncountable_compound(word):
    """Check if a compound word is likely uncountable based on its suffix."""
    for suffix in UNCOUNTABLE_SUFFIXES:
        if word.endswith(suffix) and word != suffix and len(word) > len(suffix):
            return True
    return False


def is_medical_compound(word):
    """Check if a compound word is likely a medical/anatomical term."""
    for suffix in MEDICAL_SUFFIXES:
        if word.endswith(suffix) and word != suffix and len(word) > len(suffix):
            return True
    return False

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(os.path.dirname(SCRIPT_DIR), "words.js")
DATA_DIR = os.path.join(SCRIPT_DIR, "data")


def ensure_downloaded(url, filename, desc):
    """Download a file into scripts/data/ if not already present. Returns the file path."""
    os.makedirs(DATA_DIR, exist_ok=True)
    path = os.path.join(DATA_DIR, filename)
    if os.path.exists(path):
        print(f"  {desc}: vorhanden ({path})")
    else:
        print(f"  {desc}: Lade herunter...")
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=120)
        data = resp.read()
        with open(path, "wb") as f:
            f.write(data)
        print(f"    {len(data)} bytes -> {path}")
    return path


def parse_frequency(data):
    """Parse hermitdave frequency list: 'word frequency' per line, all lowercase."""
    freq = {}
    for line in data.decode("utf-8").splitlines():
        parts = line.strip().split()
        if len(parts) == 2:
            freq[parts[0]] = int(parts[1])
    print(f"    {len(freq)} Eintraege")
    return freq


def parse_nouns(data):
    """Parse german-nouns CSV. Only entries with pos exactly 'Substantiv'.
    Returns dict: lowercase_lemma -> (capitalized_lemma, genus).
    For words with multiple genus entries, keeps the first (most common) one.
    Filters out proper nouns (Vorname, Nachname, Toponym) and plural-only lemmas."""
    nouns = {}
    proper_nouns = set()  # collect lemmas that have a proper noun entry
    reader = csv.DictReader(io.StringIO(data.decode("utf-8")))
    skipped_pos = 0
    skipped_genus = 0
    skipped_proper = 0
    skipped_duplicate = 0
    # First pass: collect all proper noun lemmas
    rows = list(reader)
    for row in rows:
        pos = row.get("pos", "").strip()
        if pos != "Substantiv" and "Substantiv" in pos:
            # e.g. "Substantiv,Vorname", "Substantiv,Toponym"
            lemma = row.get("lemma", "").strip()
            if lemma:
                proper_nouns.add(lemma.lower())
    # Second pass: collect regular nouns, excluding proper nouns
    for row in rows:
        pos = row.get("pos", "").strip()
        if pos != "Substantiv":
            if "Substantiv" in pos:
                skipped_proper += 1
            else:
                skipped_pos += 1
            continue
        lemma = row.get("lemma", "").strip()
        if not lemma or lemma.startswith("-"):
            continue
        # Skip words that also appear as proper nouns
        if lemma.lower() in proper_nouns:
            skipped_proper += 1
            continue
        genus = row.get("genus", "").strip()
        if genus not in ("m", "f", "n"):
            skipped_genus += 1
            continue
        # Keep first entry only (most common meaning in Wiktionary)
        if lemma.lower() in nouns:
            skipped_duplicate += 1
            continue
        nouns[lemma.lower()] = (lemma, genus)
    print(f"    {len(nouns)} reine Substantive ({len(proper_nouns)} Eigennamen erkannt)")
    print(f"    (uebersprungen: {skipped_pos} Nicht-Substantive, {skipped_proper} Eigennamen,")
    print(f"     {skipped_genus} ohne Genus, {skipped_duplicate} Duplikate)")
    return nouns


def parse_psycholinguistic(filepath):
    """Parse the psycholinguistic xlsx.
    Returns dict: lowercase_word -> (concreteness, valence, arousal)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Expected: Word=0, GPT_concr=5, GPT_val=6, GPT_arou=7
    col = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).strip()
        if hl in ("Word", "word", "Wort"):
            col["word"] = i
        elif hl == "GPT_concr":
            col["conc"] = i
        elif hl == "GPT_val":
            col["val"] = i
        elif hl == "GPT_arou":
            col["arou"] = i

    missing = {"word", "conc", "val", "arou"} - col.keys()
    if missing:
        print(f"    FEHLER: Spalten nicht gefunden: {missing}")
        print(f"    Verfuegbare Spalten: {headers}")
        sys.exit(1)

    print(f"    Spalten: Word={col['word']}, Concr={col['conc']}, Val={col['val']}, Arou={col['arou']}")

    data = {}
    for row in ws.iter_rows(min_row=2):
        word = row[col["word"]].value
        conc = row[col["conc"]].value
        val = row[col["val"]].value
        arou = row[col["arou"]].value
        if word and conc is not None and val is not None and arou is not None:
            try:
                data[str(word).strip().lower()] = (float(conc), float(val), float(arou))
            except (ValueError, TypeError):
                continue
    wb.close()
    print(f"    {len(data)} Bewertungen geladen")
    return data


def main():
    print("Schlummershuffle Wortlisten-Generator\n")

    # Download
    print("1. Daten herunterladen:")
    freq_path = ensure_downloaded(FREQ_URL, "de_full.txt", "Frequenzliste")
    nouns_path = ensure_downloaded(NOUNS_URL, "nouns.csv", "Substantive")
    psych_path = ensure_downloaded(CONCRETENESS_URL, "concreteness.xlsx", "Psycholinguistik")

    # Parse
    print("\n2. Daten verarbeiten:")
    print("  Frequenzliste:")
    with open(freq_path, "rb") as f:
        freq = parse_frequency(f.read())
    print("  Substantive:")
    with open(nouns_path, "rb") as f:
        nouns = parse_nouns(f.read())
    print("  Psycholinguistische Bewertungen:")
    psych = parse_psycholinguistic(psych_path)

    # Join & filter
    print(f"\n3. Zusammenfuehren und filtern:")
    print(f"  Schwellen: Konkretheit >= {CONCRETENESS_MIN}, Valenz >= {VALENCE_MIN}, Arousal <= {AROUSAL_MAX}")
    candidates = []
    stats = {"no_freq": 0, "no_psych": 0, "low_conc": 0, "low_val": 0,
             "high_arou": 0, "blocked": 0, "nom_inf": 0, "uncountable": 0,
             "medical": 0, "duplicate": 0}

    for lower, (lemma, genus) in nouns.items():
        if lower not in freq:
            stats["no_freq"] += 1
            continue
        if lower in BLOCKED:
            stats["blocked"] += 1
            continue
        if lower in DUPLICATE_REMOVE:
            stats["duplicate"] += 1
            continue
        if lower in UNCOUNTABLE or is_uncountable_compound(lower):
            stats["uncountable"] += 1
            continue
        if lower in MEDICAL or is_medical_compound(lower):
            stats["medical"] += 1
            continue
        # Nominalisierte Infinitive: Neutrum + endet auf -en (das Gehen, das Hoeren)
        # Haben verfaelschte Frequenzen weil die Verbform mitgezaehlt wird
        if genus == "n" and lower.endswith("en") and len(lower) > 3:
            stats["nom_inf"] += 1
            continue
        if lower not in psych:
            stats["no_psych"] += 1
            continue
        conc, val, arou = psych[lower]
        if conc < CONCRETENESS_MIN:
            stats["low_conc"] += 1
            continue
        if val < VALENCE_MIN:
            stats["low_val"] += 1
            continue
        if arou > AROUSAL_MAX:
            stats["high_arou"] += 1
            continue
        # Apply genus overrides
        if lower in GENUS_OVERRIDE:
            genus = GENUS_OVERRIDE[lower]
        candidates.append((freq[lower], lemma, genus, conc, val, arou))

    print(f"  Ohne Frequenzdaten:    {stats['no_freq']}")
    print(f"  Ohne Bewertungen:      {stats['no_psych']}")
    print(f"  Konkretheit zu niedrig:{stats['low_conc']}")
    print(f"  Valenz zu niedrig:     {stats['low_val']}")
    print(f"  Arousal zu hoch:       {stats['high_arou']}")
    print(f"  Blockiert:             {stats['blocked']}")
    print(f"  Unzaehlbar:            {stats['uncountable']}")
    print(f"  Medizinisch:           {stats['medical']}")
    print(f"  Duplikate:             {stats['duplicate']}")
    print(f"  Nominalisierte Inf.:   {stats['nom_inf']}")
    print(f"  -> Kandidaten:         {len(candidates)}")

    if len(candidates) < TARGET_COUNT:
        print(f"\n  WARNUNG: Nur {len(candidates)} Kandidaten, Ziel ist {TARGET_COUNT}!")
        print(f"  Eventuell Schwellenwerte anpassen.")

    # Sort by frequency, take top N
    candidates.sort(key=lambda x: x[0], reverse=True)
    final = candidates[:TARGET_COUNT]

    print(f"\n4. Top {len(final)} nach Haeufigkeit ausgewaehlt")

    # Stats
    genders = {"m": 0, "f": 0, "n": 0}
    for _, _, g, *_ in final:
        genders[g] += 1
    print(f"  Genus: m={genders['m']}, f={genders['f']}, n={genders['n']}")
    print(f"  Frequenz:    {final[0][0]:>8} – {final[-1][0]}")
    print(f"  Konkretheit: {min(c for _,_,_,c,_,_ in final):.2f} – {max(c for _,_,_,c,_,_ in final):.2f}")
    print(f"  Valenz:      {min(v for _,_,_,_,v,_ in final):.2f} – {max(v for _,_,_,_,v,_ in final):.2f}")
    print(f"  Arousal:     {min(a for _,_,_,_,_,a in final):.2f} – {max(a for _,_,_,_,_,a in final):.2f}")

    print(f"\n  Beispiele (haeufigste):")
    for fr, w, g, c, v, a in final[:20]:
        art = "ein" if g in ("m", "n") else "eine"
        print(f"    {art:4} {w:20} freq={fr:>6}  conc={c:.1f}  val={v:.1f}  arou={a:.1f}")

    # Generate words.js
    print(f"\n5. Schreibe {OUTPUT_PATH}...")
    lines = [
        "// Schlummershuffle – Deutsche Wortliste (frequenzbasiert)",
        "// Generiert von scripts/build_wordlist.py",
        "//",
        "// Quellen:",
        "//   Frequenz: hermitdave/FrequencyWords (OpenSubtitles-Korpus)",
        "//     https://github.com/hermitdave/FrequencyWords",
        "//   Substantive + Genus: gambolputty/german-nouns (Wiktionary-Extrakt)",
        "//     https://github.com/gambolputty/german-nouns",
        "//   Psycholinguistik: German Psycholinguistic Toolbox (Brysbaert et al.)",
        "//     https://osf.io/ghjd2/",
        "//",
        f"// {len(final)} konkrete, neutrale, beruhigende Substantive",
        f"// Sortiert nach Haeufigkeit im OpenSubtitles-Korpus",
        f"// Filter: Konkretheit >= {CONCRETENESS_MIN}, Valenz >= {VALENCE_MIN}, Arousal <= {AROUSAL_MAX} (Skala 1-7)",
        "//",
        "// Format: [genus, wort] – genus: m/f/n",
        "// Artikel wird zur Laufzeit abgeleitet (m/n -> 'ein', f -> 'eine')",
        "",
        "const words = [",
    ]
    for _, lemma, genus, *_ in final:
        lines.append(f"  ['{genus}','{lemma}'],")
    lines.append("];")
    lines.append("")

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"  Fertig! {len(final)} Woerter geschrieben.")


if __name__ == "__main__":
    main()
